import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from datetime import datetime
import xlsxwriter

import logging
logging.basicConfig(level=logging.INFO)


def save_tracing_table_with_pandas(profiler_table: str, output_file_name: str) -> None:
    # Convert the table string to a list of rows
    table_lines = profiler_table.strip().split('\n')
    # Skip the header separator line (contains only dashes)
    data_lines = [line for line in table_lines if not line.startswith('---') and line.strip()]

    # Parse the table into a DataFrame
    rows = []
    for line in data_lines:
        # Split by multiple spaces and filter out empty strings
        columns = [col.strip() for col in line.split('  ') if col.strip()]
        if columns:  # Only add non-empty rows
            rows.append(columns)

    # Create DataFrame (assuming first row is header)
    if rows:
        df = pd.DataFrame(rows[1:], columns=rows[0])
        
        # Save to Excel
        df.to_excel(output_file_name, index=False)
        
        logging.info(f"Profiler results saved to Excel: {output_file_name}")

def save_tracing_table_with_openpyxl(profiler_table: str, output_file_name: str) -> None:
    # Parse the table
    table_lines = profiler_table.strip().split('\n')
    data_lines = [line for line in table_lines if not line.startswith('---') and line.strip()]

    # Create workbook
    wb = Workbook()
    ws = wb.active

    # Parse and write data
    for i, line in enumerate(data_lines):
        columns = [col.strip() for col in line.split('  ') if col.strip()]
        for j, value in enumerate(columns):
            ws.cell(row=i+1, column=j+1, value=value)

    # Save to Excel
    wb.save(output_file_name)

    logging.info(f"Profiler results saved to Excel: {output_file_name}")

def save_tracing_table_with_xlsxwriter(profiler_table: str, output_file_name: str) -> None:
    # Parse the table
    table_lines = profiler_table.strip().split('\n')
    data_lines = [line for line in table_lines if not line.startswith('---') and line.strip()]

    # Create workbook
    with xlsxwriter.Workbook(output_file_name) as workbook:
        worksheet = workbook.add_worksheet()
        
        # Parse and write data
        for i, line in enumerate(data_lines):
            columns = [col.strip() for col in line.split('  ') if col.strip()]
            for j, value in enumerate(columns):
                worksheet.write(i, j, value)

    logging.info(f"Profiler results saved to Excel: {output_file_name}")