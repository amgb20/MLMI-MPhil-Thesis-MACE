import openpyxl
import os

def separate_mace_from_others(wb, source_sheet_name):
    """
    Separate MACE components from all other names in the source sheet.
    Creates two sheets: 'MACE_Only' and 'Non_MACE'.
    Works with existing workbook.
    """
    try:
        print(f"Processing sheet: {source_sheet_name}")
        source = wb[source_sheet_name]
        print(f"Found source sheet '{source_sheet_name}' with {source.max_row} rows")
        
        # Remove existing sheets if they exist
        for sheet_name in ["MACE_Only", "Non_MACE"]:
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
        
        # Create new sheets
        mace_sheet = wb.create_sheet("MACE_Only")
        non_mace_sheet = wb.create_sheet("Non_MACE")
        
        # Get header row
        header_row = next(source.iter_rows(min_row=1, max_row=1, values_only=True))
        print(f"Header row has {len(header_row)} columns")
        
        # Copy header to both sheets
        for col_idx, value in enumerate(header_row, start=1):
            mace_sheet.cell(row=1, column=col_idx, value=value)
            non_mace_sheet.cell(row=1, column=col_idx, value=value)
        
        # Initialize counters
        mace_row = 2  # start after header
        non_mace_row = 2  # start after header
        
        # Process each row and separate
        for row_idx, row in enumerate(source.iter_rows(min_row=2, values_only=True), start=2):
            cell_value = str(row[0]) if row[0] is not None else ""
            
            if "MACE/" in cell_value:
                # Add to MACE_Only sheet
                for col_idx, value in enumerate(row, start=1):
                    mace_sheet.cell(row=mace_row, column=col_idx, value=value)
                mace_row += 1
            else:
                # Add to Non_MACE sheet
                for col_idx, value in enumerate(row, start=1):
                    non_mace_sheet.cell(row=non_mace_row, column=col_idx, value=value)
                non_mace_row += 1
        
        # Print summary
        print(f"  MACE_Only sheet: {mace_row - 2} rows")
        print(f"  Non_MACE sheet: {non_mace_row - 2} rows")
        print(f"  Total processed: {(mace_row - 2) + (non_mace_row - 2)} rows")
        
    except Exception as e:
        print(f"Error during MACE separation: {str(e)}")
        raise

def create_mace_sheets_with_filtered(wb, source_sheet_name):
    """
    Create detailed MACE categorization sheets.
    Works with existing workbook.
    """
    try:
        print(f"Processing detailed MACE categorization from: {source_sheet_name}")
        source = wb[source_sheet_name]
        print(f"Found source sheet with {source.max_row} rows")
        
        # Define categories and their sheet names
        categories = {
            "node_embedding": "node_embedding",
            "radial_embedding": "radial_embedding", 
            "spherical_harmonics": "spherical_harmonics",
            "atomic_energies_fn": "atomic_energies_fn",
            "scale_shift": "scale_shift",
            "interaction": "interaction",  # MACE/Interaction[0] and MACE/Interaction[1]
            "product": "product",          # MACE/Product[0] and MACE/Product[1]
            "readout": "readout"           # MACE/Readout[0] and MACE/Readout[1]
        }
        
        # Remove existing sheets if they exist (except the original sheets)
        for sheet_name in categories.values():
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
        
        # Remove Filtered sheet if it exists
        if "Filtered" in wb.sheetnames:
            wb.remove(wb["Filtered"])
        
        # Remove Embeddings sheet if it exists
        if "Embeddings" in wb.sheetnames:
            wb.remove(wb["Embeddings"])
        
        # Create Filtered sheet for all MACE data
        filtered_sheet = wb.create_sheet("Filtered")
        
        # Create Embeddings sheet for uncategorized MACE data
        embeddings_sheet = wb.create_sheet("Embeddings")
        
        # Get header row
        header_row = next(source.iter_rows(min_row=1, max_row=1, values_only=True))
        print(f"Header row has {len(header_row)} columns")
        
        # Copy header to Filtered sheet and Embeddings sheet
        for col_idx, value in enumerate(header_row, start=1):
            filtered_sheet.cell(row=1, column=col_idx, value=value)
            embeddings_sheet.cell(row=1, column=col_idx, value=value)
        
        # Initialize counters for each category
        category_counts = {cat: 0 for cat in categories.keys()}
        category_counts["Embeddings"] = 0
        filtered_row = 2  # start after header in Filtered sheet
        embeddings_row = 2  # start after header in Embeddings sheet
        
        # Process each row and categorize
        for row_idx, row in enumerate(source.iter_rows(min_row=2, values_only=True), start=2):
            cell_value = str(row[0]) if row[0] is not None else ""
            
            if "MACE/" in cell_value:
                # Add to Filtered sheet (all MACE data)
                for col_idx, value in enumerate(row, start=1):
                    filtered_sheet.cell(row=filtered_row, column=col_idx, value=value)
                filtered_row += 1
                
                # Determine which category this row belongs to
                category_found = None
                
                if "node_embedding" in cell_value.lower():
                    category_found = "node_embedding"
                elif "radial_embedding" in cell_value.lower():
                    category_found = "radial_embedding"
                elif "spherical_harmonics" in cell_value.lower():
                    category_found = "spherical_harmonics"
                elif "atomic_energies_fn" in cell_value.lower():
                    category_found = "atomic_energies_fn"
                elif "scale_shift" in cell_value.lower():
                    category_found = "scale_shift"
                elif "interaction" in cell_value.lower():
                    category_found = "interaction"
                elif "product" in cell_value.lower():
                    category_found = "product"
                elif "readout" in cell_value.lower():
                    category_found = "readout"
                
                if category_found:
                    # Create sheet if it doesn't exist
                    if categories[category_found] not in wb.sheetnames:
                        wb.create_sheet(categories[category_found])
                    
                    target_sheet = wb[categories[category_found]]
                    
                    # Copy header if this is the first row for this category
                    if category_counts[category_found] == 0:
                        for col_idx, value in enumerate(header_row, start=1):
                            target_sheet.cell(row=1, column=col_idx, value=value)
                    
                    # Copy the row data
                    target_row = category_counts[category_found] + 2  # +2 because we start after header
                    for col_idx, value in enumerate(row, start=1):
                        target_sheet.cell(row=target_row, column=col_idx, value=value)
                    
                    category_counts[category_found] += 1
                else:
                    # This MACE data doesn't fit any specific category, add to "Embeddings"
                    for col_idx, value in enumerate(row, start=1):
                        embeddings_sheet.cell(row=embeddings_row, column=col_idx, value=value)
                    embeddings_row += 1
                    category_counts["Embeddings"] += 1
        
        # Print summary
        print(f"  Filtered sheet: {filtered_row - 2} rows (all MACE data)")
        for category, count in category_counts.items():
            sheet_name = categories[category] if category in categories else category
            print(f"  {sheet_name}: {count} rows")
        
    except Exception as e:
        print(f"Error during detailed MACE categorization: {str(e)}")
        raise

def create_comprehensive_excel_file(filename, new_filename):
    """
    Create a comprehensive Excel file with all sheets:
    - Original data (Sheet1)
    - MACE_Only and Non_MACE separation
    - Detailed MACE categorization
    """
    try:
        print(f"Loading workbook from: {filename}")
        # Load the workbook
        wb = openpyxl.load_workbook(filename)
        
        # Get source sheet (first sheet)
        if len(wb.sheetnames) == 0:
            raise ValueError("No sheets found in the workbook.")
        
        source_sheet_name = wb.sheetnames[0]
        print(f"Processing source sheet: {source_sheet_name}")
        
        # Step 1: Create MACE vs Non-MACE separation
        print("\n=== Step 1: MACE vs Non-MACE Separation ===")
        separate_mace_from_others(wb, source_sheet_name)
        
        # Step 2: Create detailed MACE categorization
        print("\n=== Step 2: Detailed MACE Categorization ===")
        create_mace_sheets_with_filtered(wb, source_sheet_name)
        
        # Print final summary
        print("\n=== Final Summary ===")
        print(f"All sheets created successfully:")
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            row_count = sheet.max_row - 1 if sheet.max_row > 1 else 0  # Subtract header
            print(f"  {sheet_name}: {row_count} data rows")
        
        # Save to new filename
        print(f"\nSaving comprehensive file to: {new_filename}")
        wb.save(new_filename)
        print(f"Comprehensive Excel file created successfully!")
        
        # Verify file was created
        if os.path.exists(new_filename):
            file_size = os.path.getsize(new_filename)
            print(f"File size: {file_size} bytes")
        else:
            print("ERROR: File was not created!")
            
    except Exception as e:
        print(f"Error during comprehensive file creation: {str(e)}")
        raise

def migrate_mace_rows(filename, new_filename):
    try:
        print(f"Loading workbook from: {filename}")
        # Load the workbook
        wb = openpyxl.load_workbook(filename)
        
        # Get source sheet (Sheet1)
        if "Sheet" not in wb.sheetnames:
            raise ValueError("Sheet not found in the workbook.")
        source = wb["Sheet"]
        print(f"Found source sheet with {source.max_row} rows")
        
        # Create new sheet (or clear if exists)
        if "Filtered" in wb.sheetnames:
            wb.remove(wb["Filtered"])
        target = wb.create_sheet("Filtered")
        
        # Copy header row (assuming first row is header)
        header_row = next(source.iter_rows(min_row=1, max_row=1, values_only=True))
        for col_idx, value in enumerate(header_row, start=1):
            target.cell(row=1, column=col_idx, value=value)
        print(f"Copied header row with {len(header_row)} columns")
        
        # Copy rows with "MACE/"
        target_row = 2  # start after header
        mace_count = 0
        for row in source.iter_rows(min_row=2, values_only=True):
            cell_value = str(row[0]) if row[0] is not None else ""
            if "MACE/" in cell_value:
                for col_idx, value in enumerate(row, start=1):
                    target.cell(row=target_row, column=col_idx, value=value)
                target_row += 1
                mace_count += 1
        
        print(f"Found and copied {mace_count} rows containing 'MACE/'")
        
        # Save to new filename only (remove the double save)
        print(f"Saving to: {new_filename}")
        wb.save(new_filename)
        print(f"Migration complete. Filtered rows written to 'Filtered' sheet in {new_filename}")
        
        # Verify file was created
        if os.path.exists(new_filename):
            file_size = os.path.getsize(new_filename)
            print(f"File successfully created with size: {file_size} bytes")
        else:
            print("ERROR: File was not created!")
            
    except Exception as e:
        print(f"Error during migration: {str(e)}")
        raise

if __name__ == "__main__":
    # Use absolute paths to avoid confusion
    current_dir = os.getcwd()
    print(f"Current working directory: {current_dir}")
    
    filename = "Experiments/numerical_stability/src/inference/results/block_level_cost/xlsx/cuda_time_total.xlsx"
    new_directory = "Experiments/numerical_stability/src/inference/results/block_level_cost/xlsx/filtered"
    
    # Check if source file exists
    if not os.path.exists(filename):
        print(f"ERROR: Source file not found: {filename}")
        exit(1)
    
    print(f"Source file exists: {filename}")
    print(f"Creating directory: {new_directory}")
    os.makedirs(new_directory, exist_ok=True)
    
    # Create comprehensive Excel file with all sheets
    comprehensive_filename = os.path.join(new_directory, "cuda_time_total_comprehensive.xlsx")
    print(f"\nCreating comprehensive Excel file: {comprehensive_filename}")
    print("This file will contain:")
    print("  - Original data (Sheet1)")
    print("  - MACE_Only and Non_MACE separation")
    print("  - Detailed MACE categorization (Filtered, interaction, product, readout, etc.)")
    
    create_comprehensive_excel_file(filename, comprehensive_filename)
    
    print(f"\nComprehensive file created successfully in: {new_directory}")
    print(f"File: {comprehensive_filename}")
